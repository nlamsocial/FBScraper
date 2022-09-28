# -*- coding: utf-8 -*-
import yaml
from yaml import Loader, SafeLoader
from enum import Enum
import time
import pandas as pd
import os
from openpyxl import load_workbook
from typing import List

LOGPATH = r"E:\DRIVE CUONG\FBSCRAP\log\logs2.xlsx"
class Action(Enum):
    NONE = 0
    APPROVE = 1
    DECLINE = 2
    BLOCK = 3
    
class PostContent():
    def __init__(self, webelement, text_content, poster = None, action = Action.NONE, keycheck = ''):
        self.webelement = webelement
        self.text_content = text_content
        self.poster = poster
        self.action = action
        self.keycheck = keycheck
        
class ApproveConfigPost():
    def __init__(self,have_keywords):
        self.have_keywords = have_keywords
        
class DeclineConfig():
    def __init__(self,livestream, reshare, have_keywords):
        self.livestream = livestream
        self.reshare = reshare
        self.have_keywords = have_keywords
        
class GroupSettingPost():
    def __init__(self, url, types, member_approve, number_of_posts, 
                 approve: ApproveConfigPost, decline: DeclineConfig):
        self.url = url
        self.types = types
        self.member_approve = member_approve
        self.number_of_posts = number_of_posts
        self.approve = approve
        self.decline = decline
        
class Profiles: 
    def __init__(self, uid, pw): 
        self.uid = uid
        self.pw = pw
        
class LogExport:
    def __init__(self, poster, content, keycheck, status=Action.NONE):
        self.time = time.strftime("%Y-%m-%d %H:%M:%S")
        self.poster = poster
        self.content = content
        self.keycheck = keycheck
        self.status = status

def save_excel_sheet(dfnew, filepath, sheetname, index=False):
    # Create file if it does not exist
    if not os.path.exists(filepath):
        dfnew.to_excel(filepath, sheet_name=sheetname, index=index)
    else:   
        with pd.ExcelWriter(filepath, engine='openpyxl', if_sheet_exists='overlay',mode='a') as writer:
            wb = load_workbook(writer, read_only=True)
            if sheetname in wb.sheetnames:
                df_read = pd.read_excel( writer, sheet_name=sheetname)
                df_total = pd.concat([dfnew, df_read], ignore_index=True)
                df_total.to_excel(writer,sheet_name = sheetname, index=False)
            else:
                dfnew.to_excel(writer, sheet_name = sheetname,index=False)
                
def update_logfile(logs: List[LogExport], sheetname):
    df = pd.DataFrame()
    for log in logs:
        dfnew = pd.DataFrame([[log.time, log.status, log.keycheck, log.poster, log.content]],
                             columns=['Time', 'Status', 'Key check','Poster', 'Content'])
        df = pd.concat([df, dfnew], ignore_index=True)
    
    save_excel_sheet(df,LOGPATH,sheetname)

def get_profiles():
    profiles = []
    with open('profiles.txt') as temp_file:
      nicks = [line.rstrip('\n') for line in temp_file]
      for nick in nicks:
          detail = nick.split("|")
          profiles.append(Profiles(detail[0],detail[1]))
    return profiles

def is_duplicate(anylist):
    if type(anylist) != 'list':
        return("Error. Passed parameter is Not a list")
    if len(anylist) != len(set(anylist)):
        return True
    else:
        return False

def get_config():
    with open('config.yaml', 'r', encoding='utf-8') as f:
        data = yaml.load(f, Loader=SafeLoader)
        
    group_settings = []
    for group_setting in data['groupmanagement']:
        url = group_setting['url']
        types = group_setting['type']
        member_approve = group_setting['member_approve']
        number_of_posts = group_setting['number_of_posts']
        approve = ApproveConfigPost(group_setting['approve_if']['have_keywords'])
        decline = DeclineConfig(group_setting['decline_if']['livestream'],
                                group_setting['decline_if']['reshare'], 
                                group_setting['decline_if']['have_keywords'])
        
        group_settings.append(GroupSettingPost(url, types, member_approve, number_of_posts, approve, decline))

    
    return group_settings